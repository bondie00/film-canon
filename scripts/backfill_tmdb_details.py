#!/usr/bin/env python3
"""
Backfill Synopsis / Runtime / Genre from TMDB for films the workbook leaves blank.

The workbook's "main data" sheet is the source of truth for these three fields
(baked in by scripts/bake_mubi_fields.py). So this script fills gaps IN THE
SHEET, not in films.json: anything written to films.json would be wiped by the
next convert-data.py run, which regenerates it from the sheet.

Two-part, mirroring merge_tmdb_images.py:
  1. FETCH  — for films missing a field AND carrying a tmdbId, call
              /movie/{id} or /tv/{id}, cache the result to data/tmdb_details.json.
              Idempotent: films already cached are skipped, so re-runs are free.
              Reads films.json for the tmdbId (merge_tmdb_images.py puts it there).
  2. APPLY  — fill blank Genre/Runtime/Synopsis cells in the sheet from the cache.
              Existing cell values are never overwritten. Runs without an API key
              (uses whatever is already cached).

Genre names are normalised on the way in (TMDB's vocabulary differs from MUBI's)
so the sheet keeps one vocabulary. That guard is why "Science Fiction" and the
two "&" television genres can no longer leak in.

This is an occasional tool, not a pipeline step. After it writes to the sheet,
re-run convert-data.py to carry the new values into films.json.

Normal build order: convert-data.py -> merge_tmdb_images.py

Setup:
  1. pip install requests
  2. Get a free API key at https://www.themoviedb.org/settings/api
  3. Set it:
        export TMDB_API_KEY=your_key_here      # macOS / Linux / Git Bash
        $env:TMDB_API_KEY = "your_key_here"   # PowerShell
  4. Run:  python scripts/backfill_tmdb_details.py
"""

import json
import os
import time
import sys
from pathlib import Path

import requests
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from genre_vocab import normalize_genres  # noqa: E402  (shared alias map)

ROOT = Path(__file__).resolve().parent.parent
FILMS_JSON = ROOT / "public" / "data" / "films.json"
XLSX = ROOT / "data" / "sight and sound data new.xlsx"
SHEET = "main data"

# Sheet column <- films.json field, for the three fields this script fills.
COLUMN_FOR_FIELD = {"synopsis": "Synopsis", "runtime": "Runtime", "genres": "Genre"}
DETAILS_JSON = ROOT / "data" / "tmdb_details.json"

TMDB_API = "https://api.themoviedb.org/3"
REQUEST_TIMEOUT = 15
SLEEP_BETWEEN = 0.05  # polite throttle; TMDB allows ~50 req/s


def needs_backfill(film):
    """Film is missing at least one of the three fields and has a TMDB match."""
    if not film.get("tmdbId"):
        return False
    return not film.get("synopsis") or not film.get("runtime") or not film.get("genres")


def load_cache():
    if DETAILS_JSON.exists():
        with DETAILS_JSON.open(encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(cache):
    tmp = DETAILS_JSON.with_suffix(".json.tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)
    tmp.replace(DETAILS_JSON)


def fetch_details(session, api_key, tmdb_id, media_type):
    """Return {synopsis, runtime, genres} from TMDB details, or None on failure."""
    mt = "tv" if media_type == "tv" else "movie"
    url = f"{TMDB_API}/{mt}/{tmdb_id}"
    resp = session.get(url, params={"api_key": api_key}, timeout=REQUEST_TIMEOUT)
    if resp.status_code == 429:
        wait = int(resp.headers.get("Retry-After", "2"))
        time.sleep(wait + 1)
        resp = session.get(url, params={"api_key": api_key}, timeout=REQUEST_TIMEOUT)
    if resp.status_code == 404:
        return {"synopsis": None, "runtime": None, "genres": []}
    resp.raise_for_status()
    d = resp.json()

    overview = (d.get("overview") or "").strip() or None

    if mt == "tv":
        ert = d.get("episode_run_time") or []
        runtime = int(ert[0]) if ert and ert[0] else None
    else:
        runtime = int(d["runtime"]) if d.get("runtime") else None

    genres = [g["name"] for g in d.get("genres", []) if g.get("name")]

    return {"synopsis": overview, "runtime": runtime, "genres": genres}


def main():
    print("Loading films.json...")
    with FILMS_JSON.open(encoding="utf-8") as f:
        films = json.load(f)

    cache = load_cache()
    todo = [f for f in films if needs_backfill(f) and str(f["key"]) not in cache]
    print(f"  {sum(needs_backfill(f) for f in films)} films missing a field with a TMDB match; "
          f"{len(cache)} already cached; {len(todo)} to fetch.")

    # ---- FETCH ----
    if todo:
        api_key = os.environ.get("TMDB_API_KEY")
        if not api_key:
            print("\n  ! TMDB_API_KEY not set — skipping fetch. Set it and re-run to pull "
                  "the missing details. (Applying whatever is already cached below.)")
        else:
            session = requests.Session()
            fetched = 0
            try:
                for i, film in enumerate(todo, 1):
                    try:
                        details = fetch_details(session, api_key, film["tmdbId"], film.get("tmdbMediaType"))
                    except requests.RequestException as e:
                        print(f"    [error] {film['FilmTitle']}: {e}")
                        continue
                    cache[str(film["key"])] = details
                    fetched += 1
                    if fetched % 25 == 0:
                        save_cache(cache)
                        print(f"    ...{fetched}/{len(todo)} fetched")
                    time.sleep(SLEEP_BETWEEN)
            finally:
                save_cache(cache)
            print(f"  Fetched {fetched} film(s); cache now has {len(cache)} entries.")

    # ---- APPLY (fill blank cells in the sheet; never overwrite) ----
    filled = {"synopsis": 0, "runtime": 0, "genres": 0}

    wb = load_workbook(XLSX)
    ws = wb[SHEET]
    header = {str(c.value): c.column for c in ws[1] if c.value is not None}
    missing_cols = [c for c in COLUMN_FOR_FIELD.values() if c not in header]
    if missing_cols:
        raise SystemExit(
            f"Sheet is missing {', '.join(missing_cols)}. "
            "Run scripts/bake_mubi_fields.py first."
        )
    key_col = header["key"]

    for row in range(2, ws.max_row + 1):
        raw = ws.cell(row=row, column=key_col).value
        if raw is None:
            continue
        entry = cache.get(str(int(raw)))
        if not entry:
            continue
        for field, column in COLUMN_FOR_FIELD.items():
            value = entry.get(field)
            if not value:
                continue
            cell = ws.cell(row=row, column=header[column])
            if cell.value not in (None, ""):
                continue
            if field == "genres":
                value = ", ".join(normalize_genres(value))
                if not value:
                    continue
            cell.value = value
            filled[field] += 1

    wb.save(XLSX)

    print(f"\n[OK] Applied to the sheet — filled "
          f"{filled['synopsis']} synopses, {filled['runtime']} runtimes, {filled['genres']} genre sets.")


if __name__ == "__main__":
    main()
