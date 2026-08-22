#!/usr/bin/env python3
"""
One-shot: bake the MUBI/TMDB-derived film fields into the workbook's "main data"
sheet, so the sheet becomes the editable source of truth for them.

WHY. Genre, runtime, synopsis and the two image URLs were re-derived on every
build by joining data/full_mubi_data.csv on mubi_id. That had three problems:

  1. The sheet ALREADY had a Genre column, but convert-data.py used it only as a
     fallback for films MUBI didn't cover -- so hand-edits to it were silently
     discarded for the 4,755 films that do have a mubi_id.
  2. That column had gone stale. 72 films still carried the genres of whatever
     film their old (wrong) mubi_id pointed at, from before the mubi_id fixes:
     Close-Up as "Comedy", Ratatouille as "Drama", an 1896 Lumiere actuality as
     "Drama, Fantasy, Thriller".
  3. full_mubi_data.csv is 20 MB, untracked, and local to one machine. Losing it
     would have stripped these fields from the next regeneration.

So we write the CURRENT, correct values (from films.json, which reflects the
corrected join plus the TMDB backfill) into the sheet, and convert-data.py then
reads them from there. After this runs, the MUBI CSV is archival.

DIRECTION MATTERS: films.json -> sheet, never the reverse. The sheet is the
stale copy.

Genres are normalised on the way in -- see GENRE_ALIASES. Four tags leaked in
from TMDB, whose vocabulary differs from MUBI's (and whose *television* list
supplied the two "&" names via Game of Thrones).

Run once:  python scripts/bake_mubi_fields.py
"""

from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from genre_vocab import normalize_genres  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
FILMS_JSON = ROOT / "public" / "data" / "films.json"
XLSX = ROOT / "data" / "sight and sound data new.xlsx"
SHEET = "main data"

# New columns, in order, inserted directly after the existing Genre column.
NEW_COLUMNS = ["Runtime", "Synopsis", "ImageUrl", "StillUrl"]


def main():
    print(f"Reading {FILMS_JSON.relative_to(ROOT)}...")
    with FILMS_JSON.open(encoding="utf-8") as f:
        films = json.load(f)

    changed_genres = 0
    by_key = {}
    for film in films:
        original = film.get("genres") or []
        genres = normalize_genres(original)
        if genres != original:
            changed_genres += 1
        by_key[int(film["key"])] = {
            "Genre": ", ".join(genres) or None,
            "Runtime": film.get("runtime"),
            "Synopsis": film.get("synopsis"),
            "ImageUrl": film.get("imageUrl"),
            "StillUrl": film.get("stillUrl"),
        }
    print(f"  {len(by_key):,} films; {changed_genres} had a genre tag normalised")

    backup = XLSX.with_suffix(".xlsx.bak")
    shutil.copy2(XLSX, backup)
    print(f"Backed up workbook to {backup.name}")

    print(f"Opening {XLSX.name} -> '{SHEET}'...")
    wb = load_workbook(XLSX)
    ws = wb[SHEET]

    header = {}
    for cell in ws[1]:
        if cell.value is not None:
            header[str(cell.value)] = cell.column
    if "key" not in header or "Genre" not in header:
        raise SystemExit("Expected 'key' and 'Genre' columns in the sheet header.")

    # Insert whichever new columns are absent, immediately after Genre, so the
    # descriptive fields sit together. Re-running is safe: existing columns are
    # reused and overwritten rather than duplicated.
    to_insert = [c for c in NEW_COLUMNS if c not in header]
    if to_insert:
        at = header["Genre"] + 1
        ws.insert_cols(at, amount=len(to_insert))
        for offset, name in enumerate(to_insert):
            ws.cell(row=1, column=at + offset, value=name)
        print(f"  inserted {len(to_insert)} column(s) after Genre: {', '.join(to_insert)}")
        header = {str(c.value): c.column for c in ws[1] if c.value is not None}
    else:
        print("  all target columns already present; overwriting in place")

    cols = {name: header[name] for name in ["Genre"] + NEW_COLUMNS}
    key_col = header["key"]

    written = 0
    missing = []
    for row in range(2, ws.max_row + 1):
        raw = ws.cell(row=row, column=key_col).value
        if raw is None:
            continue
        key = int(raw)
        values = by_key.get(key)
        if values is None:
            missing.append(key)
            continue
        for name, col in cols.items():
            # Assign to .value rather than passing value= to cell(): openpyxl
            # ignores a None passed that way, so a re-run could not clear a cell.
            ws.cell(row=row, column=col).value = values[name]
        written += 1

    if missing:
        print(f"  WARNING: {len(missing)} sheet rows had no films.json match: {missing[:10]}")

    print(f"  wrote {written:,} rows x {len(cols)} columns")
    wb.save(XLSX)
    print(f"Saved {XLSX.name}.")


if __name__ == "__main__":
    main()
